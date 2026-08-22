from __future__ import annotations

import csv
import datetime as dt
import json
from pathlib import Path

from jinja2 import Environment, select_autoescape

from .logging import get_logger
from .models import Job

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover - openpyxl is a declared dependency
    _HAS_OPENPYXL = False

_env = Environment(autoescape=select_autoescape(["html", "xml"]))
_env.trim_blocks = True

logger = get_logger("report")

REPORT_HTML = """<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<title>Job Harness Report - {{ run_ts }}</title>
<style>
body{font-family:system-ui,sans-serif;margin:24px;color:#222}
h1{font-size:1.4rem}
table{border-collapse:collapse;width:100%;font-size:0.85rem;table-layout:fixed}
th,td{border:1px solid #ddd;padding:8px;vertical-align:top;word-wrap:break-word}
th{background:#f4f4f4;text-align:left}
@page{size:A4;margin:10mm 8mm}
thead{display:table-header-group}
tr{page-break-inside:avoid}
tr.closed{background:#fff0f0}
tr.new{font-weight:bold}
a.btn{display:inline-block;padding:4px 10px;background:#1a73e8;color:#fff;text-decoration:none;border-radius:4px}
.good{color:#080}.warn{color:#b60}.bad{color:#d00}
.decision-review{background:#fef3cd;color:#856404;padding:2px 6px;border-radius:3px;font-size:12px}
.decision-auto_accept{background:#d4edda;color:#155724;padding:2px 6px;border-radius:3px;font-size:12px}
.decision-reject{background:#f8d7da;color:#721c24;padding:2px 6px;border-radius:3px;font-size:12px}
</style></head><body>
<h1>Job Harness Report - {{ run_ts }}</h1>
<p>{{ jobs|length }} jobs. {{ new_count }} genuinely new. {{ closed_count }} closed.</p>
<table><thead><tr>
<th>Fresh</th><th>Date</th><th>Title</th><th>Company</th><th>Loc</th><th>Exp</th><th>Salary</th><th>Score</th><th>Decision</th><th>Apply</th><th>Status</th><th>Domain</th><th>Sources</th>
</tr></thead><tbody>
{% for j in jobs %}
<tr class="{% if j.authentic_status=='CLOSED' %}closed{% elif j.genuinely_new %}new{% endif %}">
<td>{{ j.freshness or '—' }}</td>
<td>{{ j.date_posted }}</td>
<td>{{ j.title }}<br><small>{{ j.role }}</small></td>
<td>{{ j.company }}</td>
<td>{{ j.location or '—' }}{% if j.remote %} (remote){% endif %}</td>
<td>{{ j.experience_needed or '—' }}</td>
<td>{{ j.salary_if_present or '—' }}</td>
<td class="{% set s = j.authenticity_score %}{% if s >= 70 %}good{% elif s >= 40 %}warn{% else %}bad{% endif %}">{{ j.authenticity_score }}</td>
<td>{% if j.decision %}<span class="decision-{{ j.decision|lower }}">{{ j.decision }}{% if j.reason %}: {{ j.reason[0] }}{% endif %}</span>{% endif %}</td>
<td><a class="btn" href="{{ j.apply_url_direct }}" target="_blank">Apply</a></td>
<td class="{% if j.authentic_status=='CLOSED' %}bad{% else %}good{% endif %}">{{ j.authentic_status }}</td>
<td><small>{{ j.employer_domain }}</small></td>
<td>{{ j.seen_sources|join(', ') }}</td>
</tr>
{% endfor %}
</tbody></table></body></html>"""


def write_reports(jobs: list[Job], out_dir: str | Path, run_ts: str | None = None) -> dict[str, str | int]:
    ts = run_ts or dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    d = Path(out_dir) / ts
    d.mkdir(parents=True, exist_ok=True)

    auth = [j for j in jobs if j.authentic_status != "CLOSED"]
    new_count = sum(1 for j in jobs if j.genuinely_new and j.authentic_status != "CLOSED")
    closed_count = sum(1 for j in jobs if j.authentic_status == "CLOSED")

    json_path = d / "report.json"
    json_path.write_text(
        json.dumps([j.to_dict() for j in jobs], indent=2, default=str), encoding="utf-8"
    )

    csv_path = d / "report.csv"
    fields = [
        "freshness", "date_posted", "title", "role", "company", "location", "remote",
        "experience_needed", "salary_if_present", "seniority", "tech_stack_keywords",
        "confidence_score", "authentic_status", "employer_domain", "valid_through",
        "apply_url_direct", "source_name", "source_url", "first_seen_at",
        "genuinely_new", "seen_sources", "missing_fields", "original_url",
        "canonical_url", "final_url", "posting_id",
        "decision", "identity_score", "authenticity_score", "match_score",
        "matched_via", "possible_duplicate_of", "canonical_job_id",
        "evidence", "negative_evidence", "reason",
    ]
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(fields)
        for j in jobs:
            d_ = j.to_dict()
            w.writerow([", ".join(map(str, d_[k])) if isinstance(d_[k], list) else d_[k] for k in fields])

    html_path = d / "report.html"
    html_path.write_text(
        _env.from_string(REPORT_HTML).render(
            jobs=jobs, run_ts=ts, new_count=new_count, closed_count=closed_count
        ),
        encoding="utf-8",
    )

    xlsx_path = d / "report.xlsx"
    write_xlsx(jobs, xlsx_path, run_ts=ts, new_count=new_count, closed_count=closed_count)

    return {
        "dir": str(d),
        "json": str(json_path),
        "csv": str(csv_path),
        "html": str(html_path),
        "xlsx": str(xlsx_path),
        "pdf": "",
        "new_count": new_count,
        "closed_count": closed_count,
        "total": len(jobs),
        "authentic": len(auth),
    }


def write_pdf(html_path: str, pdf_path: str | Path) -> str:
    """Print the HTML report to a structured PDF via headless Chromium.

    Returns the PDF path on success, or '' (with a logged warning) if
    Playwright is unavailable or the conversion fails.
    """
    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        logger.warning("PDF disabled: playwright not available")
        return ""
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page()
            page.goto(Path(html_path).resolve().as_uri(), wait_until="networkidle")
            page.pdf(
                path=str(pdf_path),
                format="A4",
                print_background=True,
                margin={"top": "10mm", "bottom": "10mm", "left": "8mm", "right": "8mm"},
            )
            browser.close()
        return str(pdf_path)
    except Exception as e:
        logger.warning("PDF generation failed: %s", e)
        return ""


# ---- Excel workbook ---------------------------------------------------------

# (header, attr, width). width is a hint for column width in Excel units.
_XLSX_COLUMNS: list[tuple[str, str, int]] = [
    ("#", "row_num", 5),
    ("New", "new_flag", 6),
    ("Fresh", "freshness", 9),
    ("Posted", "date_posted", 12),
    ("Title", "title", 34),
    ("Role", "role", 26),
    ("Company", "company", 22),
    ("Location", "location", 20),
    ("Remote", "remote_flag", 8),
    ("Experience", "experience_needed", 14),
    ("Salary", "salary_if_present", 16),
    ("Seniority", "seniority", 12),
    ("Tech Stack", "tech_stack", 40),
    ("Match", "match_score", 8),
    ("Authenticity", "authenticity_score", 13),
    ("Confidence", "confidence_score", 11),
    ("Decision", "decision", 14),
    ("Reason", "reason", 40),
    ("Status", "authentic_status", 12),
    ("Source(s)", "seen_sources", 18),
    ("Domain", "employer_domain", 22),
    ("Posting ID", "posting_id", 22),
    ("Apply", "apply_short", 34),
    ("Full URL", "apply_url_direct", 46),
    ("Missing Fields", "missing_fields", 24),
    ("First Seen", "first_seen_at", 18),
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F4E79")
_SUB_FONT = Font(name="Calibri", size=11, color="333333")
_LABEL_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E79")
_BASE_FONT = Font(name="Calibri", size=11, color="222222")
_LINK_FONT = Font(name="Calibri", size=11, color="1A73E8", underline="single")
_ZEBRA_FILL = PatternFill("solid", fgColor="F4F8FC")
_NEW_FILL = PatternFill("solid", fgColor="E6F4EA")
_CLOSED_FILL = PatternFill("solid", fgColor="FCE8E6")
_THIN = Side(style="thin", color="D0D7DE")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="top")
_WRAP_CTR = Alignment(wrap_text=True, vertical="center", horizontal="center")
_TOP = Alignment(vertical="top")

_DECISION_FILLS = {
    "AUTO_ACCEPT": PatternFill("solid", fgColor="D4EDDA"),
    "REVIEW": PatternFill("solid", fgColor="FEF3CD"),
    "REJECT": PatternFill("solid", fgColor="F8D7DA"),
}


def _shorten_url_display(url: str) -> str:
    """Short, human-friendly display text for an apply hyperlink.

    Keeps the host and a tail of the path so it fits a narrow column while the
    full URL lives in its own column: 'acme.com/careers/j/1234' -> same; very
    long paths are truncated with an ellipsis in the middle.
    """
    if not url:
        return ""
    from urllib.parse import urlsplit

    parts = urlsplit(url)
    host = parts.netloc or parts.path
    path = parts.path
    if len(path) > 24:
        path = path[:14] + "…" + path[-9:]
    display = (host + path).strip("/")
    if len(display) > 46:
        display = display[:22] + "…" + display[-22:]
    return display or url


def _to_cell(v):
    if isinstance(v, (list, tuple)):
        return ", ".join(str(x) for x in v if x not in (None, "", "MISSING"))
    if v in (None, "MISSING"):
        return ""
    return v


def write_xlsx(
    jobs: list[Job],
    out_path: str | Path,
    *,
    run_ts: str | None = None,
    new_count: int = 0,
    closed_count: int = 0,
) -> str:
    """Write a designed multi-sheet Excel workbook of the run's jobs.

    Sheet 1 'Summary': run metadata, totals, decision breakdown.
    Sheet 2 'Jobs': one row per job (newest/strongest first), every field
    filled (no gaps), wrapped long text, decision color coding, zebra
    striping, frozen header + auto-filter, and a shortened apply hyperlink
    with the full URL in an adjacent column.

    Returns the path on success, or '' if openpyxl is unavailable.
    """
    if not _HAS_OPENPYXL:
        logger.warning("XLSX disabled: openpyxl not available")
        return ""
    ts = run_ts or dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    wb = Workbook()

    # ---- Summary sheet ----
    s = wb.active
    s.title = "Summary"
    s.sheet_view.showGridLines = False
    s["A1"] = "Job Harness — Run Report"
    s["A1"].font = _TITLE_FONT
    s["A2"] = f"Run: {ts}    Generated: {dt.datetime.now():%Y-%m-%d %H:%M}"
    s["A2"].font = _SUB_FONT
    rows = [
        ("Total jobs", len(jobs)),
        ("Genuinely new", new_count),
        ("Closed / removed", closed_count),
        ("Authentic (not closed)", sum(1 for j in jobs if j.authentic_status != "CLOSED")),
        ("DEGRADED (unverified link)", sum(1 for j in jobs if j.authentic_status == "DEGRADED")),
    ]
    r = 4
    for label, value in rows:
        s.cell(row=r, column=1, value=label).font = _LABEL_FONT
        s.cell(row=r, column=2, value=value).font = _BASE_FONT
        r += 1
    # decision breakdown
    decisions: dict[str, int] = {}
    for j in jobs:
        d = j.decision or "NONE"
        decisions[d] = decisions.get(d, 0) + 1
    r += 1
    s.cell(row=r, column=1, value="Decisions").font = _LABEL_FONT
    r += 1
    for d in ("AUTO_ACCEPT", "REVIEW", "REJECT", "NONE"):
        if d in decisions:
            s.cell(row=r, column=1, value=d).font = _BASE_FONT
            s.cell(row=r, column=2, value=decisions[d]).font = _BASE_FONT
            r += 1
    s.column_dimensions["A"].width = 30
    s.column_dimensions["B"].width = 14

    # ---- Jobs sheet ----
    ws = wb.create_sheet("Jobs")
    ws.sheet_view.showGridLines = False
    headers = [c[0] for c in _XLSX_COLUMNS]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP_CTR
        cell.border = _BORDER
    for ci, (_h, _a, w) in enumerate(_XLSX_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, j in enumerate(jobs, start=2):
        full_url = j.apply_url_direct or ""
        row_vals = {
            "row_num": ri - 1,
            "new_flag": "NEW" if getattr(j, "genuinely_new", False) else "",
            "freshness": j.freshness or "",
            "date_posted": j.date_posted or "",
            "title": j.title or "",
            "role": j.role or "",
            "company": j.company or "",
            "location": (j.location or "") + (" (remote)" if j.remote else ""),
            "remote_flag": "Yes" if j.remote else "No",
            "experience_needed": j.experience_needed or "",
            "salary_if_present": j.salary_if_present or "",
            "seniority": j.seniority or "",
            "tech_stack": j.tech_stack_keywords or [],
            "match_score": round(j.match_score, 1) if j.match_score else "",
            "authenticity_score": round(j.authenticity_score, 1) if j.authenticity_score else "",
            "confidence_score": j.confidence_score if j.confidence_score else "",
            "decision": j.decision or "",
            "reason": (j.reason[:3] if j.reason else []),
            "authentic_status": j.authentic_status or "",
            "seen_sources": j.seen_sources or [],
            "employer_domain": j.employer_domain or "",
            "posting_id": j.posting_id or "",
            "apply_short": _shorten_url_display(full_url),
            "apply_url_direct": full_url,
            "missing_fields": j.missing_fields or [],
            "first_seen_at": (
                dt.datetime.fromtimestamp(j.first_seen_at).strftime("%Y-%m-%d %H:%M")
                if j.first_seen_at
                else ""
            ),
        }
        for ci, (_h, attr, _w) in enumerate(_XLSX_COLUMNS, start=1):
            val = _to_cell(row_vals.get(attr, ""))
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = _BASE_FONT
            cell.border = _BORDER
            # wrapping on long-text columns; top alignment elsewhere
            if attr in ("title", "role", "tech_stack", "reason", "missing_fields",
                        "apply_short", "apply_url_direct", "seen_sources"):
                cell.alignment = _WRAP
            else:
                cell.alignment = _TOP
            # Apply hyperlink on the shortened display cell
            if attr == "apply_short" and full_url:
                cell.hyperlink = full_url
                cell.font = _LINK_FONT
        # row styling: zebra + decision tint + NEW highlight + closed red
        is_new = bool(row_vals["new_flag"])
        is_closed = j.authentic_status == "CLOSED"
        row_fill = None
        if is_closed:
            row_fill = _CLOSED_FILL
        elif is_new:
            row_fill = _NEW_FILL
        elif ri % 2 == 0:
            row_fill = _ZEBRA_FILL
        if row_fill is not None:
            for ci in range(1, len(headers) + 1):
                ws.cell(row=ri, column=ci).fill = row_fill
        # decision cell color override (wins over zebra, not over NEW/closed row)
        dci = [c[1] for c in _XLSX_COLUMNS].index("decision") + 1
        dcell = ws.cell(row=ri, column=dci)
        if j.decision in _DECISION_FILLS:
            dcell.fill = _DECISION_FILLS[j.decision]
            dcell.font = Font(name="Calibri", size=11, bold=True)
        # NEW cell bold green
        nci = [c[1] for c in _XLSX_COLUMNS].index("new_flag") + 1
        if is_new:
            ws.cell(row=ri, column=nci).font = Font(
                name="Calibri", size=11, bold=True, color="137333"
            )

    # freeze header + auto-filter; comfortable default row height
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(jobs) + 1}"
    ws.row_dimensions[1].height = 30
    for ri in range(2, len(jobs) + 2):
        ws.row_dimensions[ri].height = 42

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return str(out)
