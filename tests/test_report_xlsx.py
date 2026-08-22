from __future__ import annotations

from pathlib import Path

from jobharness.models import CLOSED, VALID_AUTHENTIC, Job
from jobharness.report import write_reports, write_xlsx

try:
    from openpyxl import load_workbook

    HAS_OPENPYXL = True
except ImportError:  # pragma: no cover
    HAS_OPENPYXL = False


def make_jobs():
    a = Job(title="Backend Engineer", company="Acme", location="Remote")
    a.role = "Backend Engineer"
    a.apply_url_direct = "https://acme.com/careers/jobs/1234567890/supercalifragilistic-long-path"
    a.authentic_status = VALID_AUTHENTIC
    a.genuinely_new = True
    a.decision = "AUTO_ACCEPT"
    a.freshness = "fresh"
    a.date_posted = "2026-08-20"
    a.experience_needed = "0-1 years"
    a.salary_if_present = "8 LPA"
    a.tech_stack_keywords = ["python", "fastapi", "docker"]
    a.match_score = 0.91
    a.authenticity_score = 88.0
    a.confidence_score = 82
    a.seen_sources = ["adzuna"]
    a.employer_domain = "acme.com"
    a.posting_id = "acme-123"
    a.reason = ["grounded salary", "ats source", "fresh posting"]
    a.missing_fields = ["valid_through"]
    a.first_seen_at = 1789000000.0
    a.compute_hash()
    a.mark_missing()

    b = Job(title="Old Role", company="Gone", location="Remote")
    b.apply_url_direct = "https://gone.com/j/2"
    b.authentic_status = CLOSED
    b.genuinely_new = False
    b.decision = "REJECT"
    b.seen_sources = ["indeed"]
    b.compute_hash()
    b.mark_missing()
    return [a, b]


def test_write_reports_produces_xlsx(tmp_path):
    if not HAS_OPENPYXL:
        import pytest

        pytest.skip("openpyxl not available")
    res = write_reports(make_jobs(), tmp_path)
    assert Path(res["xlsx"]).exists()
    wb = load_workbook(res["xlsx"])
    assert wb.sheetnames == ["Summary", "Jobs"]

    jobs_ws = wb["Jobs"]
    headers = [c.value for c in jobs_ws[1]]
    assert "Apply" in headers and "Full URL" in headers
    assert "Tech Stack" in headers and "Missing Fields" in headers

    # two data rows
    assert jobs_ws.max_row == 3


def test_xlsx_apply_link_is_shortened_hyperlink_with_full_url_column(tmp_path):
    if not HAS_OPENPYXL:
        import pytest

        pytest.skip("openpyxl not available")
    res = write_reports(make_jobs(), tmp_path)
    wb = load_workbook(res["xlsx"])
    ws = wb["Jobs"]
    header_map = {c.value: c.column for c in ws[1]}

    apply_col = header_map["Apply"]
    url_col = header_map["Full URL"]
    apply_cell = ws.cell(row=2, column=apply_col)
    url_cell = ws.cell(row=2, column=url_col)

    assert apply_cell.hyperlink is not None
    assert apply_cell.hyperlink.target == url_cell.value
    assert "acme.com" in (apply_cell.value or "")
    assert len(apply_cell.value or "") < len(url_cell.value)  # shortened display


def test_xlsx_wraps_long_text_and_fills_all_gaps(tmp_path):
    if not HAS_OPENPYXL:
        import pytest

        pytest.skip("openpyxl not available")
    res = write_reports(make_jobs(), tmp_path)
    wb = load_workbook(res["xlsx"])
    ws = wb["Jobs"]
    header_map = {c.value: c.column for c in ws[1]}

    # wrapping enabled on long-text columns
    tech = ws.cell(row=2, column=header_map["Tech Stack"])
    assert tech.alignment.wrap_text is True

    # every row covers all columns (no gaps): row 2 has salary/experience filled
    row_vals = {ws.cell(row=1, column=i).value: ws.cell(row=2, column=i).value for i in range(1, ws.max_column + 1)}
    assert row_vals["Salary"] == "8 LPA"
    assert row_vals["Experience"] == "0-1 years"
    assert row_vals["Decision"] == "AUTO_ACCEPT"
    assert row_vals["New"] == "NEW"


def test_write_xlsx_returns_empty_without_openpyxl(tmp_path, monkeypatch):
    import jobharness.report as report_mod

    monkeypatch.setattr(report_mod, "_HAS_OPENPYXL", False)
    assert write_xlsx(make_jobs(), tmp_path / "nope.xlsx") == ""
